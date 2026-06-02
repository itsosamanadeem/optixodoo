import base64
from io import BytesIO

from odoo import _, fields, models #type: ignore
from odoo.exceptions import UserError #type: ignore

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


class StockSerialImportWizard(models.TransientModel):
    _name = "stock.serial.import.wizard"
    _description = "Stock Serial Import Wizard"

    move_id = fields.Many2one("stock.move", required=True, readonly=True)
    file = fields.Binary(required=True)
    file_name = fields.Char()

    def _extract_serials_from_xlsx(self, file_content):
        if not load_workbook:
            raise UserError(_("The Python package 'openpyxl' is required for XLSX import."))

        workbook = load_workbook(filename=BytesIO(file_content), data_only=True)
        sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True)) #type: ignore
        if not rows:
            return []

        header = [str(c).strip().lower() if c else "" for c in rows[0]]
        serial_headers = {"serial", "serial number", "serial_no"}
        serial_col = next((i for i, name in enumerate(header) if name in serial_headers), 0)

        serials = []
        for row in rows[1:]:
            if not row or serial_col >= len(row):
                continue
            value = row[serial_col]
            if value is None:
                continue
            serial = str(value).strip()
            if serial:
                serials.append(serial)
        return serials

    def action_import(self):
        self.ensure_one()
        if not self.file:
            raise UserError(_("Please upload an XLSX file."))

        content = base64.b64decode(self.file)
        serials = self._extract_serials_from_xlsx(content)
        if not serials:
            raise UserError(_("No serial numbers were found in the uploaded file."))

        move = self.move_id
        existing_lines = move.move_line_ids.sorted(key=lambda l: l.id)

        idx = 0
        for line in existing_lines:
            if idx >= len(serials):
                break
            serial = serials[idx]
            line.write({"serial_no": serial})
            idx += 1

        while idx < len(serials):
            serial = serials[idx]
            self.env["stock.move.line"].create(
                {
                    "move_id": move.id,
                    "picking_id": move.picking_id.id,
                    "product_id": move.product_id.id,
                    "product_uom_id": move.product_uom.id,
                    "location_id": move.location_id.id,
                    "location_dest_id": move.location_dest_id.id,
                    "qty_done": 1.0,
                    "serial_no": serial,
                }
            )
            idx += 1

        return {"type": "ir.actions.act_window_close"}
